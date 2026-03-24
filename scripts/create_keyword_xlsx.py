#!/usr/bin/env python3
"""
Prospect Audit & Strategy — Branded Keyword Data Excel Workbook (SEO + AEO + GEO)
===================================================================================
Generates a 5-sheet branded keyword workbook with search layer classification.

Usage: python3 create_keyword_xlsx.py
Requires: openpyxl
"""

import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# ══════════════════════════════════════════════════════════════
# CONFIGURATION — Update these for each prospect
# ══════════════════════════════════════════════════════════════

PROSPECT_NAME = "{{PROSPECT_NAME}}"
PROSPECT_DOMAIN = "{{PROSPECT_DOMAIN}}"
OUTPUT_DIR = os.path.join(os.environ.get("OUTPUT_DIR", "/home/ubuntu/output"), "deliverables")
DATA_DIR = os.environ.get("OUTPUT_DIR", "/home/ubuntu/output")
LOGO_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'traffikis_logo.png')

# ── Brand Colors ──────────────────────────────────────────────
NAVY = "1B2A4A"
BLUE = "2E5090"
LIGHT_BLUE = "4A90D9"
GREEN = "7AB648"
CYAN = "00AEEF"
WHITE = "FFFFFF"
LIGHT_GREY = "F0F2F5"
MEDIUM_GREY = "E0E0E0"
DARK_GREY = "2D3748"

# Layer-specific colors
AEO_ORANGE = "E67E22"
GEO_PURPLE = "9B59B6"

# ── Styles ────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=10, color=NAVY)
subheader_fill = PatternFill(start_color=MEDIUM_GREY, end_color=MEDIUM_GREY, fill_type="solid")
data_font = Font(name="Calibri", size=10, color=DARK_GREY)
bold_data_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
title_font = Font(name="Calibri", bold=True, size=14, color=NAVY)
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
cyan_fill = PatternFill(start_color=CYAN, end_color=CYAN, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")

# Layer fill colors for conditional formatting
seo_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
aeo_fill = PatternFill(start_color="FEF3E8", end_color="FEF3E8", fill_type="solid")
geo_fill = PatternFill(start_color="F3E8FE", end_color="F3E8FE", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

# ══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════

def add_logo_and_branding(ws, title_text, num_cols=8):
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    ws["A1"].value = ""
    ws.row_dimensions[1].height = 50
    try:
        if os.path.exists(LOGO_PATH):
            img = XlImage(LOGO_PATH)
            img.width = 200
            img.height = 45
            ws.add_image(img, "A1")
        else:
            ws["A1"].value = "TrafficRadius"
            ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    except Exception:
        ws["A1"].value = "TrafficRadius"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)

    ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
    ws["A2"].fill = green_fill
    ws.row_dimensions[2].height = 4

    ws.merge_cells(f"A3:{get_column_letter(num_cols)}3")
    ws["A3"].value = title_text
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 30

    ws.merge_cells(f"A4:{get_column_letter(num_cols)}4")
    ws["A4"].fill = cyan_fill
    ws.row_dimensions[4].height = 4

    return 6

def add_footer(ws, row, num_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    ws[f"A{row}"].value = f"© Copyright 2026 TrafficRadius, All rights reserved. | Prepared for {PROSPECT_NAME} | {PROSPECT_DOMAIN}"
    ws[f"A{row}"].font = Font(name="Calibri", italic=True, size=8, color="999999")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

def write_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def write_data_row(ws, row, data, alt=False, layer=None):
    """Write a data row with optional layer-based color coding."""
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
        if layer == "SEO":
            cell.fill = seo_fill
        elif layer == "AEO":
            cell.fill = aeo_fill
        elif layer == "GEO":
            cell.fill = geo_fill
        elif alt:
            cell.fill = light_fill

def write_section_header(ws, row, text, num_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    ws[f"A{row}"].value = text
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

# ══════════════════════════════════════════════════════════════
# WORKBOOK CREATION
# ══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── SHEET 1: Dashboard ──────────────────────────────────────

ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = GREEN

start_row = add_logo_and_branding(ws, f"{PROSPECT_NAME} — Multi-Layer Keyword Opportunity Analysis", 10)

r = start_row
write_section_header(ws, r, "EXECUTIVE SUMMARY", 10)
r += 1

summary_headers = ["Metric", "Value", "Notes"]
write_header_row(ws, r, summary_headers)
r += 1

summary_data = [
    ["Total Keywords Researched", "[FROM DATA]", "All keywords validated via SEMrush"],
    ["Total Monthly Search Demand", "[FROM DATA]", "Total addressable market"],
    ["Keyword Clusters Identified", "[FROM DATA]", "Grouped by service/product category"],
    ["Average CPC", "[FROM DATA]", "Indicates commercial intent"],
    ["Keywords with Competition < 0.5", "[FROM DATA]", "Low-hanging fruit opportunities"],
    ["Top Keyword Opportunity", "[FROM DATA]", "Highest opportunity score"],
]

for i, row_data in enumerate(summary_data):
    write_data_row(ws, r, row_data, alt=(i % 2 == 1))
    r += 1

r += 1
write_section_header(ws, r, "BREAKDOWN BY SEARCH LAYER", 10)
r += 1

layer_headers = ["Search Layer", "Keywords", "Total Volume", "% of Total Volume", "Avg CPC", "Description"]
write_header_row(ws, r, layer_headers)
r += 1

layer_data = [
    ["SEO (Traditional)", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "Commercial & transactional keywords for organic rankings"],
    ["AEO (Answer Engine)", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "Question keywords for Featured Snippets & PAA"],
    ["GEO (Generative Engine)", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "[FROM DATA]", "Informational queries targeted by AI search engines"],
]

for i, row_data in enumerate(layer_data):
    layer_name = ["SEO", "AEO", "GEO"][i]
    write_data_row(ws, r, row_data, layer=layer_name)
    r += 1

r += 2
add_footer(ws, r, 10)

for col, width in zip(range(1, 11), [35, 25, 40, 15, 15, 50, 15, 15, 15, 15]):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 2: Layer Summary ─────────────────────────────────

ws_layer = wb.create_sheet("Layer Summary")
ws_layer.sheet_properties.tabColor = AEO_ORANGE

start_row = add_logo_and_branding(ws_layer, f"{PROSPECT_NAME} — Keyword Opportunity by Search Layer", 8)

r = start_row
layer_detail_headers = ["Search Layer", "Cluster", "Keywords", "Total Volume", "Avg CPC", "Avg Competition", "Traffic Value", "Priority"]
write_header_row(ws_layer, r, layer_detail_headers)
r += 1

# TODO: Replace with real data grouped by layer then cluster
# for layer in ['SEO', 'AEO', 'GEO']:
#     for cluster in clusters_in_layer:
#         write_data_row(ws_layer, r, [layer, cluster_name, ...], layer=layer)
#         r += 1

r += 2
add_footer(ws_layer, r, 8)

for col, width in zip(range(1, 9), [20, 30, 12, 15, 12, 15, 15, 10]):
    ws_layer.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 3: Cluster Summary ────────────────────────────────

ws2 = wb.create_sheet("Cluster Summary")
ws2.sheet_properties.tabColor = BLUE

start_row = add_logo_and_branding(ws2, f"{PROSPECT_NAME} — Keyword Cluster Summary", 9)

r = start_row
cluster_headers = ["Cluster", "Keywords", "SEO Kws", "AEO Kws", "GEO Kws", "Total Volume", "Avg CPC", "Traffic Value", "Priority"]
write_header_row(ws2, r, cluster_headers)
r += 1

# TODO: Replace with real cluster data from keyword_opportunities.csv
# for i, cluster_data in enumerate(clusters):
#     write_data_row(ws2, r, [cluster_data['name'], ...], alt=(i % 2 == 1))
#     r += 1

r += 2
add_footer(ws2, r, 9)

for col, width in zip(range(1, 10), [30, 12, 12, 12, 12, 15, 12, 15, 10]):
    ws2.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 4: All Keywords ───────────────────────────────────

ws3 = wb.create_sheet("All Keywords")
ws3.sheet_properties.tabColor = CYAN

start_row = add_logo_and_branding(ws3, f"{PROSPECT_NAME} — Full Keyword Universe", 7)

r = start_row
kw_headers = ["Keyword", "Search Layer", "Cluster", "Volume", "CPC", "Competition", "Opportunity Score"]
write_header_row(ws3, r, kw_headers)
r += 1

# TODO: Replace with real keyword data from keyword_opportunities.csv
# for i, kw in enumerate(all_keywords):
#     layer = kw['Search_Layer']
#     write_data_row(ws3, r, [kw['Keyword'], layer, kw['Cluster'], ...], layer=layer)
#     r += 1

r += 2
add_footer(ws3, r, 7)

for col, width in zip(range(1, 8), [40, 15, 25, 12, 10, 15, 18]):
    ws3.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 5: Top 30 Opportunities ───────────────────────────

ws4 = wb.create_sheet("Top 30 Opportunities")
ws4.sheet_properties.tabColor = GREEN

start_row = add_logo_and_branding(ws4, f"{PROSPECT_NAME} — Top 30 Keyword Opportunities", 7)

r = start_row
write_header_row(ws4, r, kw_headers)
r += 1

# TODO: Replace with top 30 keywords sorted by Opportunity Score
# for i, kw in enumerate(top_30):
#     layer = kw['Search_Layer']
#     write_data_row(ws4, r, [kw['Keyword'], layer, kw['Cluster'], ...], layer=layer)
#     r += 1

r += 2
add_footer(ws4, r, 7)

for col, width in zip(range(1, 8), [40, 15, 25, 12, 10, 15, 18]):
    ws4.column_dimensions[get_column_letter(col)].width = width

# ── Save ─────────────────────────────────────────────────────

os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, "Keyword_Strategy_Mapping.xlsx")
wb.save(output_path)
print(f"Keyword Excel saved to: {output_path}")
