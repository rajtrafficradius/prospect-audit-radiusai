#!/usr/bin/env python3
"""
Prospect Audit & Strategy — Branded 90-Day Action Plan Excel (SEO + AEO + GEO)
================================================================================
Generates a 4-sheet branded action plan workbook with tasks organized
across all three search layers.

Usage: python3 create_action_plan_xlsx.py
Requires: openpyxl
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# ══════════════════════════════════════════════════════════════
# CONFIGURATION — Update these for each prospect
# ══════════════════════════════════════════════════════════════

PROSPECT_NAME = os.environ.get("PROSPECT_NAME", "TrafficRadius Prospect")
PROSPECT_DOMAIN = os.environ.get("PROSPECT_DOMAIN", "example.com")
OUTPUT_DIR = os.path.join(os.environ.get("OUTPUT_DIR", "/home/ubuntu/output"), "deliverables")
LOGO_PATH = os.path.join(os.path.dirname(__file__), '..', 'src', 'static', 'logo.png')
DATA_DIR = os.environ.get("OUTPUT_DIR", "/home/ubuntu/output")

import json
audit_data = {}
narrative_data = {}
cro_data = {}

try:
    with open(os.path.join(DATA_DIR, "audit_findings.json"), "r") as f:
        audit_data = json.load(f)
except Exception: pass

try:
    with open(os.path.join(DATA_DIR, "strategy_narrative.json"), "r") as f:
        narrative_data = json.load(f)
except Exception: pass

try:
    with open(os.path.join(DATA_DIR, "cro_assessment.json"), "r") as f:
        cro_data = json.load(f)
except Exception: pass

# ── Brand Colors ──────────────────────────────────────────────
NAVY = "1B2A4A"
BLUE = "2E5090"
GREEN = "7AB648"
CYAN = "00AEEF"
WHITE = "FFFFFF"
LIGHT_GREY = "F0F2F5"
MEDIUM_GREY = "E0E0E0"
DARK_GREY = "2D3748"
AEO_ORANGE = "E67E22"
GEO_PURPLE = "9B59B6"

# ── Styles ────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
data_font = Font(name="Calibri", size=10, color=DARK_GREY)
bold_data_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
title_font = Font(name="Calibri", bold=True, size=14, color=NAVY)
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
cyan_fill = PatternFill(start_color=CYAN, end_color=CYAN, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
phase1_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
phase2_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
phase3_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

# Layer-specific fills
seo_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
aeo_fill = PatternFill(start_color="FEF3E8", end_color="FEF3E8", fill_type="solid")
geo_fill = PatternFill(start_color="F3E8FE", end_color="F3E8FE", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

# ══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════

def add_logo_and_branding(ws, title_text, num_cols=9):
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    ws["A1"].value = ""
    ws.row_dimensions[1].height = 50
    try:
        if os.path.exists(LOGO_PATH):
            img = XlImage(LOGO_PATH)
            img.width = 200
            img.height = 45
            ws.add_image(img, "A1")
        else:
            ws["A1"].value = "TrafficRadius"
            ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    except Exception:
        ws["A1"].value = "TrafficRadius"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)

    ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
    ws["A2"].fill = green_fill
    ws.row_dimensions[2].height = 4

    ws.merge_cells(f"A3:{get_column_letter(num_cols)}3")
    ws["A3"].value = title_text
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 30

    ws.merge_cells(f"A4:{get_column_letter(num_cols)}4")
    ws["A4"].fill = cyan_fill
    ws.row_dimensions[4].height = 4

    return 6

def add_footer(ws, row, num_cols=9):
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    ws[f"A{row}"].value = f"© Copyright 2026 TrafficRadius, All rights reserved. | Prepared for {PROSPECT_NAME} | {PROSPECT_DOMAIN}"
    ws[f"A{row}"].font = Font(name="Calibri", italic=True, size=8, color="999999")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

def write_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def write_data_row(ws, row, data, fill=None, layer=None):
    """Write a data row with optional layer-based color coding."""
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
        if layer == "SEO":
            cell.fill = seo_fill
        elif layer == "AEO":
            cell.fill = aeo_fill
        elif layer == "GEO":
            cell.fill = geo_fill
        elif fill:
            cell.fill = fill

def write_phase_header(ws, row, text, num_cols=9, fill=None):
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    ws[f"A{row}"].value = text
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=NAVY)
    ws[f"A{row}"].fill = fill if fill else light_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28

# ══════════════════════════════════════════════════════════════
# WORKBOOK CREATION
# ══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── SHEET 1: Overview ────────────────────────────────────────

ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = GREEN

start_row = add_logo_and_branding(ws, f"{PROSPECT_NAME} — 90-Day Integrated Action Plan (SEO + AEO + GEO)", 7)

r = start_row
overview_headers = ["Phase", "Focus", "Timeline", "SEO Actions", "AEO Actions", "GEO Actions", "Key Deliverables"]
write_header_row(ws, r, overview_headers)
r += 1

overview_data = [
    ("Month 1:\nFoundation",
     "Technical Foundation\n& AI Readiness",
     "Weeks 1-4",
     "Fix critical technical issues, optimize title tags & meta descriptions, set up analytics",
     "Implement FAQPage schema, add FAQ sections to top service pages, structure content for snippets",
     "Ensure AI bots not blocked, implement Organization schema, create llms.txt, verify entity signals",
     "Technical fixes complete, schema implemented, analytics live, AI accessibility confirmed"),
    ("Month 2:\nContent",
     "Content & Answer\nOptimization",
     "Weeks 5-8",
     "Publish cornerstone content for top clusters, optimize existing pages, build internal links",
     "Create how-to guides for question keywords, add Q&A structured content, optimize for PAA",
     "Publish thought leadership content, strengthen author bios, create citation-worthy resources",
     "New content published, FAQ pages live, thought leadership pieces created"),
    ("Month 3:\nAuthority",
     "Authority &\nEntity Building",
     "Weeks 9-12",
     "Launch outreach campaign, build quality backlinks, amplify top content",
     "Expand FAQ coverage, optimize for voice search, monitor snippet performance",
     "Build entity consistency across web, pursue digital PR for AI citations, strengthen brand signals",
     "Backlinks acquired, entity signals strengthened, 90-day review complete"),
]

for i, row_data in enumerate(overview_data):
    fills = [phase1_fill, phase2_fill, phase3_fill]
    write_data_row(ws, r, row_data, fill=fills[i])
    ws.row_dimensions[r].height = 80
    r += 1

r += 2
add_footer(ws, r, 7)

for col, width in zip(range(1, 8), [15, 20, 12, 40, 40, 40, 40]):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 2: Month 1 — Foundation ───────────────────────────

ws2 = wb.create_sheet("Month 1 — Foundation")
ws2.sheet_properties.tabColor = BLUE

start_row = add_logo_and_branding(ws2, f"{PROSPECT_NAME} — Month 1: Technical Foundation & AI Readiness")

r = start_row
action_headers = ["#", "Task", "Search Layer", "Pillar", "Priority", "Week", "Owner", "Status", "Notes"]
write_header_row(ws2, r, action_headers)
r += 1

month1_tasks = []
task_idx = 1
# Map SEO Quick Wins from Strategy
for qw in narrative_data.get("seo_quick_wins", []):
    month1_tasks.append([f"1.{task_idx}", qw.get("title", ""), "SEO", "Technical", "Critical", "Week 1", "Dev Team", "Not Started", qw.get("description", "")])
    task_idx += 1

# Map AEO Findings
for aeo in audit_data.get("aeo_findings", []):
    month1_tasks.append([f"1.{task_idx}", aeo.get("title", ""), "AEO", "Schema", "Important", "Week 2", "Dev Team", "Not Started", aeo.get("recommendation", "")])
    task_idx += 1

# Map GEO Findings
for geo in audit_data.get("geo_findings", []):
    month1_tasks.append([f"1.{task_idx}", geo.get("title", ""), "GEO", "Technical", "Critical", "Week 1", "Dev Team", "Not Started", geo.get("recommendation", "")])
    task_idx += 1

# Fallback
if not month1_tasks:
    month1_tasks = [
        ["1.1", "Fix robots.txt blocking issues", "SEO", "Technical", "Critical", "Week 1", "Dev Team", "Not Started", "Ensure all key pages are crawlable"],
        ["1.2", "Submit XML sitemap to Google Search Console", "SEO", "Technical", "Critical", "Week 1", "SEO Team", "Not Started", "Verify sitemap is valid and complete"],
    ]

for task in month1_tasks:
    layer = task[2]
    write_data_row(ws2, r, task, layer=layer)
    r += 1

r += 2
add_footer(ws2, r)

for col, width in zip(range(1, 10), [5, 45, 14, 12, 12, 10, 14, 14, 35]):
    ws2.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 3: Month 2 — Content ──────────────────────────────

ws3 = wb.create_sheet("Month 2 — Content")
ws3.sheet_properties.tabColor = GREEN

start_row = add_logo_and_branding(ws3, f"{PROSPECT_NAME} — Month 2: Content & Answer Optimization")

r = start_row
write_header_row(ws3, r, action_headers)
r += 1

month2_tasks = []
task_idx = 1
# Map Content Layers
for qw in narrative_data.get("content_strategy_roadmap", []):
    month2_tasks.append([f"2.{task_idx}", f"Develop Pillar: {qw.get('topic', '')}", "SEO/AEO", "Content", "Critical", "Week 5", "Content Team", "Not Started", qw.get("rationale", "")])
    task_idx += 1

# Map CRO Vulnerabilities
for cro in cro_data.get("findings", []):
    if "recommendation" in cro:
        month2_tasks.append([f"2.{task_idx}", f"CRO Fix: {cro.get('area', '')}", "ALL", "UX/CRO", "Important", "Week 6", "Dev Team", "Not Started", cro.get("recommendation", cro.get("opportunity", ""))])
    else:
        month2_tasks.append([f"2.{task_idx}", f"CRO Fix: {cro.get('area', '')}", "ALL", "UX/CRO", "Important", "Week 6", "Dev Team", "Not Started", cro.get("opportunity", "")])
    task_idx += 1

if not month2_tasks:
    month2_tasks = [
        ["2.1", "Create cornerstone content for top keyword cluster", "SEO", "Content", "Critical", "Week 5", "Content Team", "Not Started", "Target highest-volume cluster"],
        ["2.2", "Publish 2 supporting blog posts for cluster 1", "SEO", "Content", "Important", "Week 5-6", "Content Team", "Not Started", "Internal link to cornerstone"],
    ]

for task in month2_tasks:
    layer = task[2]
    write_data_row(ws3, r, task, layer=layer)
    r += 1

r += 2
add_footer(ws3, r)

for col, width in zip(range(1, 10), [5, 45, 14, 12, 12, 10, 14, 14, 35]):
    ws3.column_dimensions[get_column_letter(col)].width = width

# ── SHEET 4: Month 3 — Authority ────────────────────────────

ws4 = wb.create_sheet("Month 3 — Authority")
ws4.sheet_properties.tabColor = CYAN

start_row = add_logo_and_branding(ws4, f"{PROSPECT_NAME} — Month 3: Authority & Entity Building")

r = start_row
write_header_row(ws4, r, action_headers)
r += 1

month3_tasks = [
    # SEO tasks
    ["3.1", "Launch digital PR / outreach campaign", "SEO", "Authority", "Critical", "Week 9", "Outreach Team", "Not Started", "Target industry publications"],
    ["3.2", "Build 5-10 high-quality backlinks", "SEO", "Authority", "Important", "Week 9-12", "Outreach Team", "Not Started", "Focus on relevant, authoritative sites"],
    ["3.3", "Submit to local business directories", "SEO", "Local", "Important", "Week 9-10", "SEO Team", "Not Started", "NAP consistency across all listings"],
    ["3.4", "Amplify top-performing content on social", "SEO", "Content", "Opportunity", "Week 10", "Marketing", "Not Started", "Share cornerstone content"],
    # AEO tasks
    ["3.5", "Expand FAQ coverage to remaining service pages", "AEO", "Content", "Important", "Week 9-10", "Content Team", "Not Started", "Add FAQPage schema to all new pages"],
    ["3.6", "Optimize for voice search queries", "AEO", "Content", "Opportunity", "Week 10-11", "SEO Team", "Not Started", "Conversational, natural language answers"],
    ["3.7", "Monitor and optimize snippet performance", "AEO", "Analytics", "Important", "Week 11-12", "SEO Team", "Not Started", "Track PAA and Featured Snippet wins"],
    # GEO tasks
    ["3.8", "Build entity consistency across the web", "GEO", "Entity", "Important", "Week 9-10", "SEO Team", "Not Started", "Consistent brand info on all platforms"],
    ["3.9", "Pursue digital PR for AI citation opportunities", "GEO", "Authority", "Important", "Week 10-11", "Outreach Team", "Not Started", "Get cited in authoritative sources AI trains on"],
    ["3.10", "Strengthen brand signals on Wikipedia, Wikidata", "GEO", "Entity", "Opportunity", "Week 11", "SEO Team", "Not Started", "If eligible, create/update entries"],
    # Cross-layer
    ["3.11", "Conduct 90-day performance review (all layers)", "ALL", "Analytics", "Critical", "Week 12", "SEO Team", "Not Started", "Compare baseline vs. current across SEO, AEO, GEO"],
    ["3.12", "Prepare Month 4-6 strategy recommendations", "ALL", "Strategy", "Important", "Week 12", "SEO Team", "Not Started", "Based on 90-day learnings across all layers"],
]

for task in month3_tasks:
    layer = task[2]
    if layer == "ALL":
        write_data_row(ws4, r, task, fill=light_fill)
    else:
        write_data_row(ws4, r, task, layer=layer)
    r += 1

r += 2
add_footer(ws4, r)

for col, width in zip(range(1, 10), [5, 45, 14, 12, 12, 10, 14, 14, 35]):
    ws4.column_dimensions[get_column_letter(col)].width = width

# ── Save ─────────────────────────────────────────────────────

os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, "12_Month_Action_Plan.xlsx")
wb.save(output_path)
print(f"Action Plan Excel saved to: {output_path}")
